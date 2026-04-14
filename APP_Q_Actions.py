from flask import Flask, render_template, request, send_file, session
import os
import re
import pandas as pd
import sqlite3
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import uuid
from io import BytesIO
import json   
import pymysql

STEP1_CACHE = {}

# =========================================================
# 🔹 Flask
# =========================================================
app = Flask(__name__)
app.secret_key = "supersecret"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# =========================================================
# 🔹 Paths
# =========================================================
DB_PATH = os.path.join(BASE_DIR, "Q_Actions.db")
EXCEL_PATH = os.path.join(BASE_DIR, "Points_Feb_2026.xlsx")
UPLOAD_INFO_PATH = os.path.join(BASE_DIR, "points_upload_info.json")  # 🔥 NEW

SHEET_NAME_POINTS = "Points&Repeaters&Province"
TARGET_SHEET_DB = "All Repeaters & Affiliates"
TABLE_NAME = "repeater_actions"

_cached_df = None

# =========================================================
# 🔹 Upload Time Helpers 🔥 NEW
# =========================================================
def save_upload_time():
    data = {
        "last_upload": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    with open(UPLOAD_INFO_PATH, "w") as f:
        json.dump(data, f)

def get_upload_time():
    if os.path.exists(UPLOAD_INFO_PATH):
        with open(UPLOAD_INFO_PATH, "r") as f:
            data = json.load(f)
            return data.get("last_upload")
    return None

# =========================================================
# 🔹 Helpers
# =========================================================
def normalize_text(text):
    return re.sub(r"\s+", "", str(text)).lower().strip()

def extract_site_code(text):
    m = re.search(r"\[(.*?)\]", str(text))
    return m.group(1) if m else ""

# =========================================================
# 🔹 Points Lookup (Excel ثابت)
# =========================================================
def load_points_excel():
    global _cached_df

    if _cached_df is None:
        if not os.path.exists(EXCEL_PATH):
            raise FileNotFoundError(f"{EXCEL_PATH} not found!")

        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME_POINTS)
        df.columns = df.columns.str.strip()
        df["affiliate_normalized"] = df["Affiliate_Name"].apply(normalize_text)

        _cached_df = df

    return _cached_df


def get_repeater_and_province_from_excel(point_name):
    df = load_points_excel()
    search_value = normalize_text(point_name)

    match = df[df["affiliate_normalized"] == search_value]

    if not match.empty:
        row = match.iloc[0]
        return str(row["Repeater Class"]).strip(), str(row["Province"]).strip()

    return "No Repeater", "No Province"

# =========================================================
# 🔹 Database 
# =========================================================


# 🔥 عدل هاي حسب السيرفر مالك
DB_CONFIG = {
    "host": "localhost",      # أو IP السيرفر
    "user": "root",
    "password": "YOUR_PASSWORD",
    "database": "q_actions_db",
    "charset": "utf8mb4"
}


def connect_db():
    conn = pymysql.connect(**DB_CONFIG)
    ensure_table_exists(conn)
    return conn


def ensure_table_exists(conn):
    cursor = conn.cursor()

    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            name TEXT,
            site_code TEXT,
            q_action TEXT,
            repeater_action TEXT
        )
    """)

    conn.commit()


def excel_to_dataframe(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name=TARGET_SHEET_DB)

    df.columns = df.columns.str.strip()

    df = df.rename(columns={
        "RepeaterName / Affiliates Name": "name",
        "Site code": "site_code",
        "Q Action": "q_action",
        "Repeater Action": "repeater_action"
    })

    df = df[["name", "site_code", "q_action", "repeater_action"]]

    df["name"] = df["name"].astype(str).str.strip()
    df["site_code"] = df["site_code"].astype(str).str.strip()
    df["q_action"] = df["q_action"].fillna("No Action")
    df["repeater_action"] = df["repeater_action"].fillna("No Action")

    if df.empty:
        raise ValueError("Empty Excel")

    return df


# 🔥 بديل to_sql
def save_to_db(df):
    conn = connect_db()
    cursor = conn.cursor()

    # حذف البيانات القديمة
    cursor.execute(f"DELETE FROM {TABLE_NAME}")

    # إدخال البيانات الجديدة
    for _, row in df.iterrows():
        cursor.execute(f"""
            INSERT INTO {TABLE_NAME} (name, site_code, q_action, repeater_action)
            VALUES (%s, %s, %s, %s)
        """, (
            row["name"],
            row["site_code"],
            row["q_action"],
            row["repeater_action"]
        ))

    conn.commit()
    conn.close()


def get_actions_from_db(cursor, value):
    normalized_value = normalize_text(value)

    cursor.execute(f"""
        SELECT q_action, repeater_action
        FROM {TABLE_NAME}
        WHERE REPLACE(LOWER(name), ' ', '') = %s
        LIMIT 1
    """, (normalized_value,))

    row = cursor.fetchone()

    if row:
        return row[0] or "No Action", row[1] or "No Action"

    return "No Action", "No Action"


def get_q_action_by_site_code(cursor, site_code):
    normalized_value = normalize_text(site_code)

    cursor.execute(f"""
        SELECT q_action
        FROM {TABLE_NAME}
        WHERE REPLACE(LOWER(site_code), ' ', '') = %s
        LIMIT 1
    """, (normalized_value,))

    row = cursor.fetchone()

    if row and row[0] and row[0].strip().lower() != "no action":
        return row[0].strip()

    return None

# =========================================================
# 🔹 Core Logic
# =========================================================
def get_actions(points_text):
    results = []
    conn = connect_db()
    cursor = conn.cursor()

    points_list = [p.strip() for p in points_text.splitlines() if p.strip()]

    for idx, point in enumerate(points_list):
        repeater, province = get_repeater_and_province_from_excel(point)

        special_note = province.lower() in ["basrah", "najaf"]

        if repeater == "No Repeater":
            results.append({
                "id": idx,
                "point": point,
                "repeater": "",
                "province": "",
                "q_action": None,
                "r_action": None,
                "not_found": True,
                "special_note": False
            })
            continue

        q_db, r_db = get_actions_from_db(cursor, point)

        if (q_db.lower() != "no action") or (r_db.lower() != "no action"):
            q_action, r_action = q_db, r_db
        else:
            site_code = extract_site_code(repeater)

            if site_code:
                r_site = get_q_action_by_site_code(cursor, site_code)

                if r_site:
                    q_action, r_action = None, r_site
                else:
                    q_action, r_action = None, "Default => (+80%) International & Local CDN 100%"
            else:
                q_action, r_action = None, "Default => (+80%) International & Local CDN 100%"

        results.append({
            "id": idx,
            "point": point,
            "repeater": repeater,
            "province": province,
            "q_action": q_action,
            "r_action": r_action,
            "not_found": False,
            "special_note": special_note,
            "no_action_needed": False
        })

    conn.close()
    return results




def apply_rule_local(step1_results, rule):
    rows = []

    step1_results = [r for r in step1_results if not r.get("not_found")]
    step1_results = sorted(step1_results, key=lambda x: x["id"])

    for idx, item in enumerate(step1_results, start=1):
        site_code = extract_site_code(item["repeater"])
        rule_to_apply = item.get("special_rule") or rule

        q_action = str(item.get("q_action", "")).lower()
        action_type = "Add" if q_action == "no action" else "Edit"

        rows.append([
            idx,
            item["point"],
            item["repeater"],
            site_code,
            "Affiliates",
            item["province"],
            action_type,
            rule_to_apply
        ])

    df = pd.DataFrame(rows, columns=[
        "#",
        "RepeaterName / Affiliates Name",
        "RepeaterName / Connected from",
        "Site code",
        "Traffic Area",
        "DataCenter",
        "Action Type",
        "Action needed on Repeater / Affiliates"
    ])

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for i, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            if i == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output


# =========================================================
# 🔹 NEW ROUTE (Upload Points Excel)
# =========================================================
@app.route("/upload_points_excel", methods=["POST"])
def upload_points_excel():
    global _cached_df

    file = request.files.get("points_file")

    if not file or not file.filename.endswith(".xlsx"):
        return "❌ Invalid file", 400

    try:
        _cached_df = None  # clear cache
        file.save(EXCEL_PATH)  # replace file
        save_upload_time()  # save time
        return "✅ Points file updated", 200
    except Exception as e:
        return str(e), 500

# =========================================================
# 🔹 Routes (بدون تغيير)
# =========================================================
@app.route("/set_special_rule/<int:point_id>", methods=["POST"])
def set_special_rule(point_id):
    data = request.get_json() or {}
    rule = data.get("rule")

    key = session.get("data_key")
    results = STEP1_CACHE.get(key, [])

    for r in results:
        if r.get("id") == point_id:
            r["special_rule"] = rule
            break

    session["step1_results"] = results
    STEP1_CACHE[key] = results

    return "", 204


@app.route("/mark_no_action/<int:point_id>", methods=["POST"])
def mark_no_action(point_id):
    data = request.get_json(silent=True) or {}
    undo = data.get("undo", False)

    key = session.get("data_key")
    results = STEP1_CACHE.get(key, [])

    for r in results:
        if r.get("id") == point_id:
            r["no_action_needed"] = not undo
            break

    session["step1_results"] = results
    STEP1_CACHE[key] = results

    return "", 204


@app.route("/", methods=["GET", "POST"])
def index():
    message = None
    excel_file = None
    point_result = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "update_db":
            file = request.files.get("excel_file")

            if file and file.filename.endswith(".xlsx"):
                try:
                    df = excel_to_dataframe(file)
                    save_to_db(df)
                    message = "✅ DB updated"
                    excel_file = file.filename
                except Exception as e:
                    message = str(e)
            else:
                message = "❌ Upload valid Excel"

        elif action == "get_actions":
            points = request.form.get("points")
            if points:
                point_result = get_actions(points)

                key = str(uuid.uuid4())
                STEP1_CACHE[key] = point_result

                session["data_key"] = key
                session["total_points"] = len(point_result)

                session["show_special_note"] = any(
                    r["province"].lower() in ["najaf","basrah"]
                    for r in point_result if not r["not_found"]
                )

        elif action == "apply_rule":
            key = session.get("data_key")
            data = STEP1_CACHE.get(key)
            rule = request.form.get("rule")

            if data and rule:
                filtered = [r for r in data if not r.get("not_found") and not r.get("no_action_needed")]

                if filtered:
                    file = apply_rule_local(filtered, rule)
                    return send_file(
                        file,
                        as_attachment=True,
                        download_name="Sales_order.xlsx",
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    message = "⚠️ No valid points"

    return render_template(
        "WEB_Q_Action.html",
        message=message,
        excel_file=excel_file,
        point_result=point_result,
        show_special_note=session.get("show_special_note", False),
        total_points=session.get("total_points", 0),
        step1_results=session.get("step1_results", []),
        last_upload_time=get_upload_time()  # 🔥 NEW
    )

# =========================================================
# 🔹 Run
# =========================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
