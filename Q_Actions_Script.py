import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from Q_Action_Database import (
    connect_db,
    get_actions_from_db,
    get_q_action_by_site_code
)

from Points_Lookup import get_repeater_and_province_from_excel

def normalize_text(text):
    return re.sub(r"\s+", "", str(text)).strip().lower()

def extract_site_code(text):
    m = re.search(r"\[(.*?)\]", str(text))
    return m.group(1) if m else ""

def get_actions(points_text):
    results = []
    conn = connect_db()
    cursor = conn.cursor()

    points_list = [p.strip() for p in points_text.splitlines() if p.strip()]

    for point in points_list:
        repeater, province = get_repeater_and_province_from_excel(point)

        if repeater == "No Repeater":
            results.append({
                "point": point,
                "repeater": "",
                "province": "",
                "q_action": "",
                "r_action": "",
                "q_action_repeater": "",
                "not_found": True
            })
            continue

        q_action, r_action = get_actions_from_db(cursor, point)
        site_code = extract_site_code(repeater)
        q_action_repeater = get_q_action_by_site_code(cursor, site_code)

        results.append({
            "point": point,
            "repeater": repeater,
            "province": province,
            "q_action": q_action,
            "r_action": r_action,
            "q_action_repeater": q_action_repeater,
            "not_found": False
        })

    conn.close()
    return results

def apply_rule_local(step1_results, rule):
    import os
    import pandas as pd
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from Q_Actions_Script import extract_site_code

    rows = []
    # استبعد النقاط غير الموجودة
    step1_results = [r for r in step1_results if not r.get("not_found")]

    # ✅ بناء صفوف الاكسل مع تحديد Action Type حسب Q Action
    for idx, item in enumerate(step1_results, start=1):
        point = item["point"]
        repeater_full = item["repeater"]
        province = item["province"]
        site_code = extract_site_code(repeater_full)

        # إذا موجود special_rule استعمله، وإلا الرول العام
        rule_to_apply = item.get("special_rule", rule)

        # تحديد Action Type حسب Q Action
        q_action = str(item.get("q_action", "")).strip().lower()
        if q_action == "no action":
            action_type = "Add"
        else:
            action_type = "Edit"

        rows.append([
            idx,
            point,
            repeater_full,
            site_code,
            "Affiliates",
            province,
            action_type,  # العمود الجديد بعد التعديل
            rule_to_apply
        ])

    # إنشاء DataFrame
    df = pd.DataFrame(rows, columns=[
        "#",
        "RepeaterName / Affiliates Name",
        "RepeaterName / Connected from",
        "Site code",
        "Traffic Area",
        "DataCenter",
        "Action Type",
        "Action needed on Repeater / Affiliates"
    ])

    # تحديد اسم الملف مع التاريخ الحالي
    today_str = datetime.now().strftime("%d-%m-%Y")
    filename = f"Sales order {today_str}.xlsx"

    # مسار المجلد داخل المشروع
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    full_path = os.path.join(output_dir, filename)

    # حفظ ملف الاكسل
    df.to_excel(full_path, index=False)

    # تنسيق الاكسل
    wb = load_workbook(full_path)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if row_idx == 1:
                cell.font = Font(size=12, bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

    # ضبط عرض الأعمدة
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(full_path)
    return full_path